import io
import json
import re
import uuid
from collections import Counter
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from flask import Blueprint, request, jsonify, send_file

from .claude_client import call_claude
from .session_store import get_xml, strip_bpmndi
from .rag.pipeline import run as rag_run

tests_bp = Blueprint("tests", __name__)

# ─────────────────────────────────────────────
# STATIC SYSTEM PROMPT  (cached after first call)
# ─────────────────────────────────────────────
_TEST_SYSTEM = """BPMN QA expert. Analyze the BPMN XML and generate comprehensive test cases covering all execution paths.
Return ONLY a valid JSON array — no markdown, no explanation.

Each object must have exactly these keys:
  id (TC-001 sequential) | suite | name (≤60 chars) | description | path (ordered array of exact BPMN element names from XML) | preconditions | steps (numbered, \\n-separated string) | expected_result | test_type

suite ∈ ["Happy Path","Gateway Branches","Boundary Events","Exception Handling","Negative Tests"]
test_type: "Positive" | "Negative"

Generate 8–20 tests covering:
- 1 Happy Path (full start→end flow)
- Every exclusiveGateway branch (≥1 test each)
- Parallel paths if parallelGateway present
- All boundary events (timer/error/message) if present
- Exception/error flows
- ≥2 Negative tests (invalid input, missing data, unauthorized)

Rules: use ONLY element names from the XML; paths must be valid execution flows; do not invent elements."""


# ─────────────────────────────────────────────
# PROMPT BUILDER  (user message — dynamic only)
# ─────────────────────────────────────────────

def build_test_prompt(xml: str, rag_context: str = "") -> str:
    parts = []
    if rag_context:
        parts.append(f"QA PATTERNS:\n{rag_context}")
    # Strip BPMNDi (visual coords) before sending — irrelevant for test generation.
    # Removes ~40-60% of XML, saving hundreds of input tokens per call.
    parts.append(f"BPMN:\n{strip_bpmndi(xml)}")
    return "\n\n".join(parts)


# ─────────────────────────────────────────────
# ROUTES
# ─────────────────────────────────────────────

@tests_bp.route('/generate-tests', methods=['POST'])
def generate_tests():
    data       = request.get_json(silent=True) or {}
    session_id = data.get('session_id', '')
    xml        = data.get('xml', '').strip()

    if not xml and session_id:
        xml = get_xml(session_id) or ''
    if not xml:
        return jsonify({"error": "No BPMN XML provided"})

    rag_context, rag_meta = rag_run(xml, context_type="test")

    raw = call_claude(
        prompt=build_test_prompt(xml, rag_context),
        system=_TEST_SYSTEM,
        max_tokens=4000,
    )
    if not raw:
        return jsonify({"error": "Failed to reach Claude API"})

    try:
        raw   = re.sub(r'```json\s*', '', raw)
        raw   = re.sub(r'```\s*',     '', raw).strip()
        start = raw.find('[')
        end   = raw.rfind(']') + 1
        if start == -1 or end == 0:
            return jsonify({"error": "No JSON array found in response", "raw": raw[:300]})
        test_cases = json.loads(raw[start:end])
        return jsonify({"success": True, "test_cases": test_cases, "rag": rag_meta})
    except json.JSONDecodeError as e:
        return jsonify({"error": f"Failed to parse test cases: {e}", "raw": raw[:300]})


@tests_bp.route('/export-tests', methods=['POST'])
def export_tests():
    data         = request.get_json(silent=True) or {}
    test_cases   = data.get('test_cases', [])
    process_name = data.get('process_name', 'BPMN Process')

    if not test_cases:
        return jsonify({"error": "No test cases provided"})

    wb = openpyxl.Workbook()

    purple_fill = PatternFill("solid", fgColor="FF667EEA")
    green_fill  = PatternFill("solid", fgColor="FF28A745")
    white_bold  = Font(color="FFFFFF", bold=True, size=11)
    bold        = Font(bold=True)
    status_fills = {
        'Pass':    PatternFill("solid", fgColor="FFD4EDDA"),
        'Fail':    PatternFill("solid", fgColor="FFF8D7DA"),
        'Blocked': PatternFill("solid", fgColor="FFFFF3CD"),
        'Not Run': PatternFill("solid", fgColor="FFF8F9FA"),
    }
    alt_fill = PatternFill("solid", fgColor="FFF2F3FF")

    # ── Test Cases Sheet ───────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Test Cases"
    headers = ['TC ID', 'Suite', 'Test Name', 'Type', 'Description',
               'Path', 'Preconditions', 'Steps', 'Expected Result',
               'Status', 'Notes', 'Executed By', 'Date']

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill      = purple_fill
        cell.font      = white_bold
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 32

    for row_idx, tc in enumerate(test_cases, 2):
        status    = tc.get('status', 'Not Run')
        path_val  = tc.get('path', [])
        path_str  = ' → '.join(path_val) if isinstance(path_val, list) else str(path_val or '')
        steps_str = (tc.get('steps', '') or '').replace('\\n', '\n')

        values = [
            tc.get('id', ''),            tc.get('suite', ''),
            tc.get('name', ''),          tc.get('test_type', ''),
            tc.get('description', ''),   path_str,
            tc.get('preconditions', ''), steps_str,
            tc.get('expected_result', ''), status,
            tc.get('notes', ''),         tc.get('executed_by', ''),
            tc.get('date', ''),
        ]

        max_lines = max((len(str(v).split('\n')) for v in values if v), default=1)
        ws.row_dimensions[row_idx].height = max(18, min(max_lines * 15, 200))

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=str(val) if val is not None else '')
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            if col_idx == 10:
                cell.fill = status_fills.get(status, status_fills['Not Run'])
            elif row_idx % 2 == 0:
                cell.fill = alt_fill

    col_widths = [10, 20, 30, 12, 42, 35, 32, 52, 42, 12, 30, 18, 14]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = 'A2'

    # ── Summary Sheet ──────────────────────────────────────────────────────
    ws_sum = wb.create_sheet("Summary")
    ws_sum['A1'] = "BPMN Process Test Report"
    ws_sum['A1'].font = Font(bold=True, size=16)
    ws_sum['A2'] = f"Process: {process_name}"
    ws_sum['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws_sum['A4'] = f"Total Test Cases: {len(test_cases)}"
    ws_sum['A4'].font = bold

    for ref, val in [('A6', 'Suite'), ('B6', 'Count')]:
        ws_sum[ref] = val
        ws_sum[ref].fill = green_fill
        ws_sum[ref].font = white_bold
        ws_sum[ref].alignment = Alignment(horizontal='center')

    row = 7
    for suite, count in Counter(tc.get('suite', 'Unknown') for tc in test_cases).items():
        ws_sum[f'A{row}'] = suite
        ws_sum[f'B{row}'] = count
        row += 1

    row += 1
    for ref, val in [(f'A{row}', 'Status'), (f'B{row}', 'Count')]:
        ws_sum[ref] = val
        ws_sum[ref].fill = purple_fill
        ws_sum[ref].font = white_bold
        ws_sum[ref].alignment = Alignment(horizontal='center')

    row += 1
    for status, count in Counter(tc.get('status', 'Not Run') for tc in test_cases).items():
        ws_sum[f'A{row}'] = status
        ws_sum[f'B{row}'] = count
        if fill := status_fills.get(status):
            ws_sum[f'A{row}'].fill = fill
            ws_sum[f'B{row}'].fill = fill
        row += 1

    ws_sum.column_dimensions['A'].width = 30
    ws_sum.column_dimensions['B'].width = 12
    wb.active = 0

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f"test_cases_{uuid.uuid4().hex[:6]}.xlsx",
    )
